import os
import platform
import subprocess

dirpath = os.path.dirname(os.path.realpath(__file__))

#Running scripts based on os in order to install missing libraries
if platform.system() == "Windows":
    subprocess.run([os.path.join(dirpath, 'ModuleInstallWindows.bat')])
elif platform.system() == "Darwin":
    subprocess.call(['sh', os.path.join(dirpath, 'ModuleInstallMac.sh')])
elif platform.system() == "Linux":
    subprocess.call(['sh', os.path.join(dirpath, 'ModuleInstallLinux.sh')])

import openpyxl
from openpyxl.styles import PatternFill

import tkinter
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk

#Universal codon table, set up as a dictionary with codons as keys for amino acids
codonTable = {
    "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L", "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S",
    "TAT": "Y", "TAC": "Y", "TAA": "STOP", "TAG": "STOP", "TGT": "C", "TGC": "C", "TGA": "STOP", "TGG": "W",
    "CTT": "L", "CTC": "L", "CTA": "L", "CTG": "L", "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P",
    "CAT": "H", "CAC": "H", "CAA": "Q", "CAG": "Q", "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R",
    "ATT": "I", "ATC": "I", "ATA": "I", "ATG": "M", "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T",
    "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K", "AGT": "S", "AGC": "S", "AGA": "R", "AGG": "R",
    "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V", "GCT": "A", "GCC": "A", "GCA": "A", "GCG": "A",
    "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E", "GGT": "G", "GGC": "G", "GGA": "G", "GGG": "G",
}

outputPath = os.path.join(dirpath, 'largeOutput.txt')
mamBacPath = os.path.join(dirpath, 'mambac.xlsx')
codonatedPath = os.path.join(dirpath, 'inputsequence_Codonated.xlsx')
mamBac = openpyxl.load_workbook(mamBacPath)
largeOutput = open(outputPath, "w")
codonated = openpyxl.Workbook()

#Initializing some variables that need to be tracked outside of the functions
outputRows = {}
inputFilenames = {}
patterns = []
sequences = []
itterrations = [1]
previousFilename = [""]
previousPairs = []

#Opens the file interface for selecting a sequence input file and imports all the data from the file if it is a new file
def inputFileSelect():
    filename = filedialog.askopenfilename(initialdir = "/", title = "Select a File", filetypes = (('text files', '*.txt'), ('All files', '*.*')))

    #Tracks the name of the input files for reference for large output
    if filename not in inputFilenames.keys():
        inputFilenames.update({filename:itterrations[0]})
        itterrations[0]+=1

    if previousFilename[0] != filename:
        sequences.clear()
        
        inputFileLabel.configure(text="Selected File: "+filename)
        
        inputFile = open(filename, "r")

        for line in inputFile:
            if not line.isspace():
                sequences.append(line)
        
        previousFilename[0] = filename

        inputFile.close()

#Opens file select interface for pattern template file, gets all input from the file and adds it to the listbox display and the pattern list
def templateFileSelect():
    filename = filedialog.askopenfilename(initialdir = "/", title = "Select a File", filetypes = (('text files', '*.txt'), ('All files', '*.*')))

    templateFileLabel.configure(text="Selected File: "+filename)

    patternTemplate = open(filename, "r")

    tempPatterns = patternTemplate.readlines()

    for pattern in tempPatterns:
        add(pattern)

    patternTemplate.close()

#Clears all items from the pattern listbox
def clearPatterns():
    patterns.clear()
    patternList.delete(0, END)

#Deletes selected items from the pattern listbox
def deleteSelected():
    selections = patternList.curselection()
    loops = 0

    for i in selections:
        temp = i - loops
        patterns.remove(patternList.get(temp))
        patternList.delete(temp)
        loops+=1

#Grabs patterns from the text entry box
def add(temp):
    temp = temp.strip()

    isInt = True
    invalid = False

    try:
        int(temp)
    except ValueError:
        isInt = False

    #If the entered value is an int and not already in the list of patterns,
    #adds it to the list of patterns, and updates the list of patterns in the UI
    if ((isInt is not False) and (temp not in patterns)):
        for i in temp:
            if i not in ['1', '2', '3', '4']:
                invalid = True

        if invalid is False:
            patterns.append(temp)
            patternList.insert(END, temp)

    patternEntry.delete(0, 'end')

#Cleans up sequence input from the input file
def cleanSequence(sequence, j):

    sequence = sequence.strip()

    for i in range(0, len(sequence)):
        if sequence[i] not in codonTable.values():
            messagebox.showerror('Input Error', 'Error: Sequence #' + str(j+1) + ' is not a valid sequence and will not be translated, invalid character at index ' + str(i))
            return "error"

    return sequence

#Actual work of translating the sequence and generating output
def translate(aminoAcidSequence, translationSheet, patternSeparated, translatedSequence, sequence, sequenceType, output, pattern, patternTracker, translatedConcat):

    for x in range(0, len(aminoAcidSequence)):

        #Looping through the first row of the translation sheet, containing each amino acid
        for i in range(1, translationSheet.max_column + 1):

            #Checks if the amino acid in the current cell matches the current one in the sequence, and if it does, begins translating using that column
            if (translationSheet.cell(1, i)).value == aminoAcidSequence[x]:
                temp = patternSeparated[patternTracker] + 1

                #Checks if that particular digit in the pattern is viable for that amino acid (i.e. W, which only has 1 codon translation, so a pattern of 2, 3, or 4 wouldn't work)
                #Continues reducing temp until it finds a row that isn't empty
                while ((translationSheet.cell(temp, i)).value is None):
                    temp -= 1

                translatedSequence.append((translationSheet.cell(temp, i)).value)
                translatedConcat = translatedConcat + (translationSheet.cell(temp, i)).value

                if len(sequence) <= 32000:
                    if ((sequenceType == 'amino' and len(sequence) < 10000) or sequenceType == 'nucleic'):
                        (output.cell(outputRows[pattern], x + 1)).value = aminoAcidSequence[x]
                        (output.cell(outputRows[pattern] + 1, x + 1)).value = translatedSequence[x]

                        #Highlights cells where replacements were needed in red
                        if (temp != patternSeparated[patternTracker] + 1):
                            (output.cell(outputRows[pattern] + 1, x + 1)).fill = PatternFill(patternType='solid', fgColor='ff0000')
                
                break

        if (patternTracker == len(patternSeparated)-1):
            patternTracker = 0
        else:
            patternTracker += 1

    return translatedConcat

#Primary function, runs through all the patterns and sequences and generates the output
def codonate(patterns, mamBac, codonated, codonTable, outputRows, largeOutput, sequences):
    #print(sequences)
    #print(patterns)

    #Note - could add and if or exception here that cuts the function short if no sheet has been selected in order to prevent issues
    sheetName = sheetList.get()
    translationSheet = mamBac[sheetName]

    overlapChars = ['T', 'A', 'G', 'C']

    #print(patterns)

    for pattern in patterns:
        patternSeparated = []

        #Used for tracking the row position for each pattern sheet in the output file, ensures that all patterns are in the dictionary only once
        if pattern not in outputRows.keys():
            outputRows.update({pattern: 1})

        #print(outputRows)

        #Creating the sheet for the pattern in the output file if it doesn't exist and setting it as the active sheet for output
        if pattern not in codonated.sheetnames:
            codonated.create_sheet(pattern)
        output = codonated[pattern]

        #print(codonated.sheetnames)

        for i in range(0, len(pattern)):
            patternSeparated.append(int(pattern[i]))

        #Loop to run through all the sequences from the input file
        for j in range(0, len(sequences)):
            aminoAcidSequence = []
            translatedSequence = []
            translatedConcat = ""
            patternTracker = 0
            sequenceType = "nucleic"

            #print(patternSeparated)

            sequence = cleanSequence(sequences[j], j)

            if sequence != "error":
                if (pattern, sequence) not in previousPairs:
                    previousPairs.append((pattern, sequence))

                    #Decides to the best of the program's ability whether the sequence is a codon sequence or an amino acid sequence
                    #First checks if it is not divisible by 3, as if it isn't, its impossible for it to be a codon sequence
                    if (len(sequence) % 3 != 0):
                        sequenceType = "amino"
                    #Then checks if any character in the sequence is not a character that overlaps between amino acids and codons (as codons use the letters T, A, C, and G, which can also be amino acids)
                    #If any 1 character is not an overlap, then it's guaranteed to be an amino acid sequence
                    else:
                        for i in sequence:
                            if (i not in overlapChars):
                                sequenceType = "amino"
                                break
                    #Very unlikely off chance that someone could have an "amino acid sequence" of say TAGCTA, which could technically be an amino acid sequence, and it would be treated as a codon sequence

                    #Note - Excel files have a cell character limit of ~32000 and a column limit of ~16000, so any codon sequence > 32000 characters will not be able to put it's concatonated ouput in an excel file
                    #The same goes for any amino sequence >10000 characters, as the concatonated translated output will be 3 times that (3 character/codon, 1 amino acid:1 codon)
                    #On top of that, both risk hitting the column limit when outputting the seperated form of the amino acid sequence or translated sequence, thus requiring any sequences past those lengths be output to a text file with a different format

                    #Ensures when updating the spreadsheet that the current sequence isnt an oversized sequence, as those will be saved elsewhere
                    if len(sequence) <= 32000:
                        if ((sequenceType == 'amino' and len(sequence) < 10000) or sequenceType == 'nucleic'):
                            (output.cell(outputRows[pattern], 1)).value = sheetName
                            outputRows[pattern] += 1

                    #print(sequenceType)

                    if sequenceType == "nucleic":
                        if len(sequence) <= 32000:
                            (output.cell(outputRows[pattern], 1)).value = sequence
                            outputRows[pattern] += 1
                        for i in range(0, len(sequence), 3):
                            aminoAcidSequence.append(codonTable[sequence[i:i+3]])
                    else:
                        for i in sequence:
                            aminoAcidSequence.append(i)

                    #print(aminoAcidSequence)

                    #Core operations of the function, outputs concatenated string of translated sequence
                    translatedConcat = translate(aminoAcidSequence, translationSheet, patternSeparated, translatedSequence, sequence, sequenceType, output, pattern, patternTracker, translatedConcat)

                    if (len(translatedConcat) > 32000):
                        largeOutput.write(sheetName + ', ' + pattern + ', File #' + inputFilenames.get(previousFilename) + ', Sequence #' + str(j+1) + ': ' + translatedConcat + '\n\n\n\n')
                    else:
                        (output.cell(outputRows[pattern] + 2, 1)).value = translatedConcat
                        outputRows[pattern] += 4

                #print(translatedSequence)
                #print(translatedConcat)

    messagebox.showinfo(title=None, message="Complete, close this popup to return to program")

#All UI elements being created and organized
root = Tk()
root.title("Codonator")
root.geometry('500x400')

inputSelect = Button(root, text="Select Input File", command=lambda: inputFileSelect())
inputSelect.place(relx=0.3, y=30, anchor=CENTER)

inputFileLabel = Label(root, text="Selected File: ")
inputFileLabel.place(relx=0.3, y=60, anchor=CENTER)

sheetLbl = Label(root, text="Select translation sheet by species: ")
sheetLbl.place(relx=0.3, y=100, anchor=CENTER)

sheetList = ttk.Combobox(root, width=25)
sheetList['values'] = mamBac.sheetnames[0: len(mamBac.sheetnames) - 1]
sheetList.place(relx=0.3, y=130, anchor=CENTER)
sheetList.current()

patternLbl = Label(root, text="Enter patterns, hit 'add' to add a pattern: ")
patternLbl.place(relx=0.3, y=170, anchor=CENTER)

patternEntry = Entry(root, width=25)
patternEntry.place(relx=0.3, y=190, anchor=CENTER)

addButton = Button(root, text="Add", command=lambda: add(patternEntry.get()))
addButton.place(relx=0.3, y=220, anchor=CENTER)

templateSelect = Button(root, text="Select Pattern Template", command=lambda: templateFileSelect())
templateSelect.place(relx=0.3, y=270, anchor=CENTER)

templateFileLabel = Label(root, text="Selected File: ")
templateFileLabel.place(relx=0.3, y=300, anchor=CENTER)

runButton = Button(root, text="Run patterns", command=lambda: codonate(patterns, mamBac, codonated, codonTable, outputRows, largeOutput, sequences))
runButton.place(relx=0.3, y=350, anchor=CENTER)

patternListLbl = Label(root, text="Current patterns: ")
patternListLbl.place(relx=0.66, y=12)

patternList = Listbox(root, width=25, height=12, selectmode=MULTIPLE)
patternList.place(relx=0.6, y=38)

scroll_bar = Scrollbar(root)
scroll_bar.pack(side=RIGHT, fill=Y)
scroll_bar.config(command=patternList.yview)

clearButton = Button(root, text="Clear patterns", command=lambda: clearPatterns())
clearButton.place(relx=0.67, y = 240)

deleteSelections = Button(root, text="Delete Selected Patterns", command=lambda: deleteSelected())
deleteSelections.place(relx=0.61, y=270)

root.mainloop()

#Prevents an IndexError when saving the output excel sheet after closing the UI window early, only clears the default sheet if other sheets have been added to the file (i.e. patterns)
if len(codonated.sheetnames) > 1:
    del codonated['Sheet']

largeOutput.close()
codonated.save(codonatedPath)


